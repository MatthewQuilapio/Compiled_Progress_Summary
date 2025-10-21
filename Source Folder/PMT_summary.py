"""
Author: Engr. Angelo Matthew Cortez Quilapio
Date: 2025-02-26
Version: 8.1
Description: This is use to compile all Dave_Summary_Project and paste it into PMT.
"""
import os
import subprocess
import tkinter as tk
from datetime import datetime
import shutil
import pandas as pd
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import PatternFill
from pathlib import Path

root = tk.Tk()
root.title("PMT")
root.geometry('300x200')
root.iconbitmap('mascot.ico')
root.configure(bg = 'indigo')
root.resizable(False, False)
global initial_copy_range
#frame = tk.Frame(root, width=50, height=50, bg='red')
#frame.grid()
fill_NG = PatternFill(patternType="solid", fgColor="3BB2BF", bgColor="A227D3")
fill_OK = PatternFill(patternType="solid", fgColor="A227D3", bgColor="3BB2BF")
def main_input_directory():
    my_label1 = tk.Label(root, text="PMT Directory", bg = 'yellow', anchor='w', width=10, fg = 'black')
    my_label1.grid(row=0, column=0, ipadx=5, ipady=0)
    pmt_dir = tk.Entry(root)
    pmt_dir.grid(row=0, column=1, padx=5, pady=5)

    my_label2 = tk.Label(root, text="Summary CSVs", bg = 'yellow', anchor='w', width=10, fg = 'black')
    my_label2.grid(row=1, column=0, ipadx=5, ipady=0)
    csv_dir = tk.Entry(root)
    csv_dir.grid(row=1, column=1, padx=5, pady=5)

    button1 = tk.Button(root, text="Confirm Target", command = lambda : search_csv(csv_dir), bg = 'red')
    button1.grid(row=4, column=1, padx=5, pady=5)

    button2 = tk.Button(root, text="Browse", command = lambda : get_local_pmt_dir(pmt_dir), bg = 'lightgreen')
    button2.grid(row=0, column=2, padx=5, pady=5)
    
    button3 = tk.Button(root, text="Copy to Local PMT", command = lambda : copy_local_pmt(pmt_dir), bg = 'lightgreen')
    button3.grid(row=3, column=1, padx=5, pady=5)

    button4 = tk.Button(root, text="Browse", command = lambda : get_csv_folders(csv_dir), bg = 'red')
    button4.grid(row=1, column=2, padx=5, pady=5)

def float_or_not(cell_value):
    try:
        return float(cell_value)
    except ValueError:
        return None

def int_or_not(cell_value):
    try:
        return int(cell_value)
    except ValueError:
        return None

def create_xlsx_files(root_folder, ctr, fld_id):
    for csv_file in next(os.walk(root_folder))[2]:
            if ".csv" in csv_file:
                #print(csv_file)
                csv_directory = os.path.join(root_folder, csv_file).replace("\\", "/")
                #print(csv_directory)
                if (fld_id[0] in csv_directory):
                    df_new_file = pd.read_csv(csv_directory)
                    str_new_file_name = 'D:/temp_csv/' + 'file_OK' + str(ctr) + '.xlsx'
                    GFG = pd.ExcelWriter(str_new_file_name)
                    df_new_file['Execution Date'] = pd.to_datetime(df_new_file['Execution Date']).dt.strftime('%Y/%m/%d %H:%M').replace('/0', '/').replace(' 0', ' ')
                    df_new_file['Start Date'] = pd.to_datetime(df_new_file['Start Date']).dt.strftime('%Y/%m/%d %H:%M').replace('/0', '/').replace(' 0', ' ')
                    df_new_file['End Date'] = pd.to_datetime(df_new_file['End Date']).dt.strftime('%Y/%m/%d %H:%M').replace('/0', '/').replace(' 0', ' ')
                    df_new_file.to_excel(GFG, index=False)
                    GFG.close()
                elif (fld_id[1] in csv_directory):
                    df_new_file = pd.read_csv(csv_directory)
                    str_new_file_name = 'D:/temp_csv/' + 'file_NG' + str(ctr) + '.xlsx'
                    df_new_file['Execution Date'] = pd.to_datetime(df_new_file['Execution Date']).dt.strftime('%Y/%m/%d %H:%M').replace('/0', '/').replace(' 0', ' ')
                    df_new_file['Start Date'] = pd.to_datetime(df_new_file['Start Date']).dt.strftime('%Y/%m/%d %H:%M').replace('/0', '/').replace(' 0', ' ')
                    df_new_file['End Date'] = pd.to_datetime(df_new_file['End Date']).dt.strftime('%Y/%m/%d %H:%M').replace('/0', '/').replace(' 0', ' ')
                    GFG = pd.ExcelWriter(str_new_file_name)
                    df_new_file.to_excel(GFG, index=False)
                    GFG.close()
                elif (fld_id[2] in csv_directory):
                    df_new_file = pd.read_csv(csv_directory)
                    str_new_file_name = 'D:/temp_csv/' + 'file_Pa' + str(ctr) + '.xlsx'
                    df_new_file['Execution Date'] = pd.to_datetime(df_new_file['Execution Date']).dt.strftime('%Y/%m/%d %H:%M').replace('/0', '/').replace(' 0', ' ')
                    df_new_file['Start Date'] = pd.to_datetime(df_new_file['Start Date']).dt.strftime('%Y/%m/%d %H:%M').replace('/0', '/').replace(' 0', ' ')
                    df_new_file['End Date'] = pd.to_datetime(df_new_file['End Date']).dt.strftime('%Y/%m/%d %H:%M').replace('/0', '/').replace(' 0', ' ')
                    GFG = pd.ExcelWriter(str_new_file_name)
                    df_new_file.to_excel(GFG, index=False)
                    GFG.close()
                ctr = ctr + 1

def printing(xlsx_dir, fail_sheet, ws, wb, cell_array, initial_copy_range):
        data = pd.read_excel(xlsx_dir)
        row_range = data.shape[0]
        #print("RRange")
        #print(row_range)
        #print(row_range)
        # Strip whitespace from column names
        data.columns = data.columns.str.strip()

        #print(data.columns)
        for x in range(row_range):
            range_ctr = 0
            print("Loading <(°ロ°)> <(°ロ°)> <(°ロ°)>")
            #print("X")
            #print(x)
            for range_ctr in range(13):
                #print(range_ctr)
                #print("\n")
                #print(initial_copy_range)
                if (data['Status'].iloc[x]=="OK" or data['Status'].iloc[x]=="NG"):
                    cell_val = data['MILS Coverage - Reactis'].iloc[x]
                    MILS_value = float_or_not(cell_val)
                    if (MILS_value != None):
                        if((MILS_value <= 100.0) and (MILS_value >= 0.0)):
                            cell_val_2 = data['SILS Coverage - Reactis for C'].iloc[x]
                            SILS_value = float_or_not(cell_val_2)
                            if((SILS_value <= 100.0) and (SILS_value >= 0.0)):
                                cell_val_3 = data['SILS Coverage - Reactis for C'].iloc[x]
                                SILS_value_MC = float_or_not(cell_val_3)
                                if((SILS_value_MC <= 100) and (SILS_value_MC >= 0)):
                                    ws[str(chr(65 + range_ctr))+str(initial_copy_range)] = data[cell_array[range_ctr]].iloc[x]
                                    ws[str(chr(65 + range_ctr))+str(initial_copy_range)].alignment = Alignment(horizontal="general", vertical="center")
                                else:
                                    fail_sheet[str(chr(65 + range_ctr))+str(initial_copy_range)] = data[cell_array[range_ctr]].iloc[x]
                                    fail_sheet[str(chr(65 + range_ctr))+str(initial_copy_range)].alignment = Alignment(horizontal="general", vertical="center")
                            else:
                                fail_sheet[str(chr(65 + range_ctr))+str(initial_copy_range)] = data[cell_array[range_ctr]].iloc[x]
                                fail_sheet[str(chr(65 + range_ctr))+str(initial_copy_range)].alignment = Alignment(horizontal="general", vertical="center")
                        else:                          
                                    #messagebox.showerror("Μήνυμα Σφάλματος", "Υπάρχει ένα σφάλμα στο MILS")
                            fail_sheet[str(chr(65 + range_ctr))+str(initial_copy_range)] = data[cell_array[range_ctr]].iloc[x]
                            fail_sheet[str(chr(65 + range_ctr))+str(initial_copy_range)].alignment = Alignment(horizontal="general", vertical="center")
                    else:
                                #messagebox.showerror("Μήνυμα Σφάλματος", "Υπάρχει ένα σφάλμα στο MILS")
                        fail_sheet[str(chr(65 + range_ctr))+str(initial_copy_range)] = data[cell_array[range_ctr]].iloc[x]
                        fail_sheet[str(chr(65 + range_ctr))+str(initial_copy_range)].alignment = Alignment(horizontal="general", vertical="center")
                else:
                            #messagebox.showerror("錯誤信息", "狀態有錯誤")
                    fail_sheet[str(chr(65 + range_ctr))+str(initial_copy_range)] = data[cell_array[range_ctr]].iloc[x]
                    fail_sheet[str(chr(65 + range_ctr))+str(initial_copy_range)].alignment = Alignment(horizontal="general", vertical="center")
                            #ws[str(chr(65 + range_ctr))+str(initial_copy_range)] = data[cell_array[range_ctr]].iloc[x]
                            #ws[str(chr(65 + range_ctr))+str(initial_copy_range)].alignment = Alignment(horizontal="left")      
                       
                wb.save('PMT.xlsx')
                #print(range_ctr)
                #print(initial_copy_range)
                range_ctr = range_ctr + 1
            cellC = ws[str(chr(67))+str(initial_copy_range)]
            cellE = ws[str(chr(69))+str(initial_copy_range)]
            cellF = ws[str(chr(70))+str(initial_copy_range)]
            cellG = ws[str(chr(71))+str(initial_copy_range)]
            cellL = ws[str(chr(76))+str(initial_copy_range)]
            cellM = ws[str(chr(77))+str(initial_copy_range)]
            cellC.alignment = Alignment(horizontal="right", vertical="center")
            cellE.alignment = Alignment(horizontal="general", vertical="center")
            cellF.alignment = Alignment(horizontal="general", vertical="center")
            cellG.alignment = Alignment(horizontal="general", vertical="center")
            cellL.alignment = Alignment(horizontal="right", vertical="center")
            cellM.alignment = Alignment(horizontal="right", vertical="center")
            cellC.number_format = 'yyyy/m/d'
            cellL.number_format = 'yyyy/m/d'
            cellM.number_format = 'yyyy/m/d'
            initial_copy_range = initial_copy_range + 1
        return initial_copy_range

#for deleting blank spaces
def remove_blank(max_sheet, work_sheet):
    for y in range(max_sheet):
    #print(ws[str(chr(65)) + str(y + 1)].value)
    #ws.delete_rows(y)
    #print(max_non_fail_sheet - y)
        if (work_sheet[str(chr(65)) + str(max_sheet - y)].value) != None:
            #print("Pass")
            #print(work_sheet[str(chr(65)) + str(max_sheet - y)].value)
            pass
        else:
            #print("Fail")
            #print(work_sheet[str(chr(65)) + str(max_sheet - y)].value)
            work_sheet.delete_rows(max_sheet - y)

def initialize_header(work_sheet):
    work_sheet['A1'] = 'Model Name'
    work_sheet['B1'] = 'Subsystem Name'
    work_sheet['C1'] = 'Execution Date'
    work_sheet['D1'] = 'Status'
    work_sheet['E1'] = 'MILS Coverage - Reactis'
    work_sheet['F1'] = 'SILS Coverage - Reactis for C'
    work_sheet['G1'] = 'SILS Coverage - MC-Verifier'
    work_sheet['H1'] = 'Model File Path'
    work_sheet['I1'] = 'Subsystem'
    work_sheet['J1'] = 'C Source Files'
    work_sheet['K1'] = 'Result Folder Path'
    work_sheet['L1'] = 'Start Date'
    work_sheet['M1'] = 'End Date'


def search_csv(csv_dir):
    wb = Workbook()
    ws = wb.active
    
    initialize_header(ws)

    ws.title = 'OK_Sheet'
    fail_sheet = wb.create_sheet('Failure', 1)
    initialize_header(fail_sheet)
    pattern_sheet = wb.create_sheet('Pattern_Sheet', 1)
    initialize_header(pattern_sheet)
    ng_sheet = wb.create_sheet('NG_Sheet', 1)
    initialize_header(ng_sheet)
    
    wb.save('PMT.xlsx')
    

    fld_id = ['01_Dave実行結果(OK)', '02_Dave実行結果(NG)', '03_パターン手修正']
    folders_csv = []
    for root, dirs, files in os.walk(csv_dir.get().replace("\\", "/")):
        #print("root")
        #print(root)
        folders_csv.append(root.replace("\\", "/"))

        #print(folders_csv[0])
        """print("Current directory:", root.replace("\\", "/"))
        print("Subdirectories:", dirs"
        print("Files:", files)"""
    ctr = 0
    ctr_2 = 0
    #print("root")
    root_1 = folders_csv[1]
    root_2 = folders_csv[2]
    root_3 = folders_csv[3]
    #print(root_1)
    #all_files_1 = [next(os.walk(root_1))[2]]
    #print(all_files_1)
    create_xlsx_files(root_1, ctr, fld_id)
    create_xlsx_files(root_2, ctr, fld_id)
    create_xlsx_files(root_3, ctr, fld_id)

    
        #destination_path = "D:/PythonforSBRNG/Scan_File/py/py/PMT/PMT.l..xlsx"
        #dest_workbook = Workbook.load_workbook(filename=destination_path)
    cell_array = ['Model Name',
                    'Subsystem Name',
                    'Execution Date',
                    'Status',
                    'MILS Coverage - Reactis',
                    'SILS Coverage - Reactis for C',
                    'SILS Coverage - MC-Verifier',
                    'Model File Path',
                    'Subsystem',
                    'C Source Files',
                    'Result Folder Path',
                    'Start Date',
                    'End Date',
                    ]
    xlsx_files = next(os.walk('D:/temp_csv/'))[2]
    #print("XXX")
    #print(xlsx_files)
    
        #copy from all csv summary reports->xlsx the data inside into PMT.l..xlsx
    initial_copy_range = 2 
    for xlsx_file in xlsx_files: 
        temp_xlsx_directory = os.path.join('D:/temp_csv/', xlsx_file).replace("\\", "/")
            #print(temp_xlsx_directory)
        #print("Dir:")
        if ('file_OK' in temp_xlsx_directory):
            initial_copy_range_1 = printing(temp_xlsx_directory, fail_sheet, ws, wb, cell_array, initial_copy_range)
            initial_copy_range = initial_copy_range_1
            #printing(temp_xlsx_directory, fail_sheet, ws, wb, cell_array, initial_copy_range)
            #initial_copy_range = initial_copy_range + 1
    
    initial_copy_range = 2  
    for xlsx_file in xlsx_files:
        temp_xlsx_directory = os.path.join('D:/temp_csv/', xlsx_file).replace("\\", "/")
        if ('file_Pa' in temp_xlsx_directory):
            initial_copy_range_2 = printing(temp_xlsx_directory, fail_sheet, pattern_sheet, wb, cell_array, initial_copy_range)
            initial_copy_range = initial_copy_range_2
            #printing(temp_xlsx_directory, fail_sheet, pattern_sheet, wb, cell_array, initial_copy_range)
            #initial_copy_range = initial_copy_range + 1
    
    initial_copy_range = 2  
    for xlsx_file in xlsx_files:
        temp_xlsx_directory = os.path.join('D:/temp_csv/', xlsx_file).replace("\\", "/")
        if ('file_NG' in temp_xlsx_directory):
            initial_copy_range_3 = printing(temp_xlsx_directory, fail_sheet, ng_sheet, wb, cell_array, initial_copy_range)
            initial_copy_range = initial_copy_range_3
            #printing(temp_xlsx_directory, fail_sheet, ng_sheet, wb, cell_array, initial_copy_range)
            #initial_copy_range = initial_copy_range + 1

        #for deletion of temporary csv
    for xlsx_file_delete in xlsx_files:
        del_xlsx_directory = os.path.join('D:/temp_csv/', xlsx_file_delete).replace("\\", "/")
        os.remove(del_xlsx_directory)
    

    
    print("Done Copying")
    #print(folders_csv)

    max_non_fail_sheet = wb['OK_Sheet'].max_row
    max_ng_sheet = wb['NG_Sheet'].max_row
    max_pattern_sheet = wb['Pattern_Sheet'].max_row
    max_fail_sheet = wb['GeneticFailure'].max_row
    remove_blank(max_non_fail_sheet, ws)
    remove_blank(max_ng_sheet, ng_sheet)
    remove_blank(max_pattern_sheet, pattern_sheet)
    remove_blank(max_fail_sheet, fail_sheet)
    #print(max_non_fail_sheet)


    #ws.delete_rows(146)
    wb.save('PMT.xlsx')       
    


#if tool will be shared to dtph, remove function for copying sheet from PMT.l..xlsx to Official PMT file, never share the code without knowing reason baka mawalan ng malalagay sa portfolio

#this is for the textbox input
def get_local_pmt_dir(pmt_dir):
    pmtdir = filedialog.askopenfilename(title="Select a File", filetypes=[("Excel", "*.xlsx"), ("All Files", "*.*")])
    pmt_dir.delete(0, tk.END)
    pmt_dir.insert(0, pmtdir)

#this is for the textbox input
def get_csv_folders(csv_dir):
    csvdir = filedialog.askdirectory(title="Select a Folder")
    csv_dir.delete(0, tk.END)
    csv_dir.insert(0, csvdir)

#this function is to copy 
def copy_local_pmt(pmt_dir):
    pmtdir = pmt_dir.get().replace("\\", "/")   
    pmt_workbook = load_workbook(pmtdir)
    src_workbook = load_workbook('PMT.xlsx')
    #pmt_workbook.remove(pmt_workbook["OK_Sheet"])
    #pmt_workbook.remove(pmt_workbook["NG_Sheet"])
    #pmt_workbook.remove(pmt_workbook["Pattern_Sheet"])
    #if 
    src_sheet_1 = src_workbook['OK_Sheet']
    src_sheet_2 = src_workbook['NG_Sheet']
    src_sheet_3 = src_workbook['Pattern_Sheet']
    #des_sheet_3 = pmt_workbook.create_sheet("Pattern_Sheet", -1)
    #des_sheet_2 = pmt_workbook.create_sheet("NG_Sheet", -1)
    des_sheet_1 = pmt_workbook.create_sheet("Compiled_Sheet", -1)
    max_row_1 = src_sheet_1.max_row
    max_column_1 = src_sheet_1.max_column
    max_row_2 = src_sheet_2.max_row
    max_column_2 = src_sheet_2.max_column
    max_row_3 = src_sheet_3.max_row
    max_column_3 = src_sheet_3.max_column
    
    for row in range(1, max_row_1 + 1):
        for col in range(1, max_column_2 + 1): 
            des_sheet_1.cell(row=row, column=col).value = src_sheet_1.cell(row=row, column=col).value
    #print("Max Row and Col 1")
    #print(max_row_1)
    #print(max_column_1)
    row_ini = 2
    for row in range(max_row_1 + 1, max_row_1 + max_row_2):
        for col in range(1, max_column_2 + 1):
            des_sheet_1.cell(row=row, column=col).value = src_sheet_2.cell(row=row_ini, column=col).value
        row_ini = row_ini + 1
    row_ini = 2     
    for row in range(max_row_1 + max_row_2, max_row_1 + max_row_2 + max_row_3):
        for col in range(1, max_column_3 + 1):
            des_sheet_1.cell(row=row, column=col).value = src_sheet_3.cell(row=row_ini, column=col).value
        row_ini = row_ini + 1
    print("Max Row and Col 3")
    print(max_row_3)
    print(max_column_3)
            #des_sheet_1.cell(row=row, column=col).alignment = Alignment(horizontal="general", vertical="general")
    """pmt_workbook.remove(pmt_workbook["For_copy"])
    des_sheet = pmt_workbook.create_sheet("For_copy", -1)
    max_row = src_sheet.max_row
    max_column = src_sheet.max_column
    for row in range(1, max_row + 1):
        for col in range(1, max_column + 1):
            des_sheet.cell(row=row, column=col).value = src_sheet.cell(row=row, column=col).value
            des_sheet.cell(row=row, column=col).alignment = Alignment(horizontal="left")
    dave_project_summary_sheet = pmt_workbook['DaveProject_Summary']
    for_copy_sheet = pmt_workbook['For_copy']
    #print(pmt_workbook['For_copy'].max_row)
    #print(dave_project_summary_sheet[str(chr(67)) + str(2)].value)
    for function_row in range (for_copy_sheet.max_row):
        #print(function_row)
        for search_function in range (dave_project_summary_sheet.max_row):
            if (dave_project_summary_sheet[str(chr(67)) + str(search_function + 2)].value) == (for_copy_sheet[str(chr(66)) + str(function_row + 2)].value):
                #for_copy_sheet[str(chr(66)) + str(function_row + 1)].fill = fill_OK
                if ((dave_project_summary_sheet[str(chr(67)) + str(search_function + 2)].value) != None) or ((for_copy_sheet[str(chr(66)) + str(function_row + 2)].value) != None):
                    content = str(dave_project_summary_sheet[str(chr(67)) + str(search_function + 2)].value) + " and " + str(for_copy_sheet[str(chr(66)) + str(function_row + 2)].value)

                    # Define the file name
                    file_name = "OK_File.txt"

                    # Open the file in write mode
                    with open(file_name, "w") as file:
                        # Write the content to the file
                        file.write(content)
                        file.write('\n')
            else:
                #for_copy_sheet[str(chr(66)) + str(function_row + 1)].fill = fill_NG
                if ((dave_project_summary_sheet[str(chr(67)) + str(search_function + 2)].value) != None) or ((for_copy_sheet[str(chr(66)) + str(function_row + 2)].value) != None):
                    content = str(dave_project_summary_sheet[str(chr(67)) + str(search_function + 2)].value) + " and " + str(for_copy_sheet[str(chr(66)) + str(function_row + 2)].value)

                    # Define the file name
                    file_name = "NG_File.txt"

                    # Open the file in write mode
                    with open(file_name, "w") as file:
                        # Write the content to the file
                        file.write(content)
                        file.write('\n')"""
    pmt_workbook.save(pmtdir)
    pmt_workbook.close()
    
    


def main_function():
    main_input_directory()
    root.mainloop()

main_function()
